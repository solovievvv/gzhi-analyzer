"""
Чтение xlsx/xls и извлечение транзакций по листам.

Поддерживаемые форматы:
  - .xlsx, .xlsm  — через openpyxl
  - .xls          — пробуем как xlsx (часто переименованные xlsx),
                    затем через xlrd (настоящий бинарный xls)
"""
from typing import Optional
import io
import re
import openpyxl

from app.core.models import (
    TransactionRow, SheetResult, AnalysisResult, FolderResult
)
from app.core.filters import is_deposit_transfer
from app.core.date_parser import parse_date
from app.core.deduplicator import deduplicate_folder


# ── Загрузка файла ────────────────────────────────────────────────────────────

def _load_workbook(filepath: str):
    """
    Открывает файл и возвращает openpyxl Workbook.
    Определяет формат по содержимому (сигнатуре), а не по расширению.
    """
    with open(filepath, "rb") as f:
        raw = f.read()

    if not raw:
        raise RuntimeError("файл пустой (0 байт)")

    # PK = ZIP-архив = xlsx/xlsm (даже если расширение .xls)
    if raw[:2] == b"PK":
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # D0 CF = OLE2 = настоящий бинарный .xls (BIFF5/8 в OLE-контейнере)
    if raw[:2] == b"\xd0\xcf":
        return _xls_via_xlrd(raw)

    # BOF 0x09xx = старый BIFF без OLE-контейнера (Excel 2.0–4.0).
    # openpyxl такие не открывает, xlrd — читает. CODEPAGE часто нет → cp1251.
    if raw[:1] == b"\x09" and raw[1:2] in (b"\x00", b"\x02", b"\x04", b"\x08"):
        return _xls_via_xlrd(raw, encoding_override="cp1251")

    # Последняя попытка — просто openpyxl
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


def _xls_via_xlrd(raw: bytes, encoding_override: Optional[str] = None):
    """Открывает бинарный/BIFF .xls через xlrd и конвертирует в openpyxl."""
    try:
        import xlrd
    except ImportError:
        raise RuntimeError(
            "Файл в формате .xls (старый Excel). Установите xlrd: pip install xlrd"
        )
    kwargs = {"file_contents": raw}
    if encoding_override:
        kwargs["encoding_override"] = encoding_override
    return _xlrd_to_openpyxl(xlrd.open_workbook(**kwargs))


def _xlrd_to_openpyxl(xls_wb):
    """
    Конвертирует xlrd Workbook в openpyxl Workbook в памяти.
    Сохраняет значения ячеек и имена листов.
    """
    import xlrd
    from datetime import datetime, date as date_type
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_idx in range(xls_wb.nsheets):
        xls_ws = xls_wb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=xls_ws.name)

        for r in range(xls_ws.nrows):
            for c in range(xls_ws.ncols):
                cell = xls_ws.cell(r, c)
                ctype = cell.ctype

                if ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                elif ctype == xlrd.XL_CELL_TEXT:
                    value = cell.value
                elif ctype == xlrd.XL_CELL_NUMBER:
                    value = cell.value
                    # Целые числа без дробной части
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                elif ctype == xlrd.XL_CELL_DATE:
                    try:
                        tup = xlrd.xldate_as_tuple(cell.value, xls_wb.datemode)
                        value = datetime(*tup) if tup[3:] != (0, 0, 0) else date_type(*tup[:3])
                    except Exception:
                        value = cell.value
                elif ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                else:
                    value = cell.value

                # openpyxl отклоняет управляющие символы — чистим (иначе
                # IllegalCharacterError и весь файл не читается).
                if isinstance(value, str):
                    value = ILLEGAL_CHARACTERS_RE.sub("", value)

                ws.cell(row=r + 1, column=c + 1, value=value)

    return wb


# ── Поиск заголовков ─────────────────────────────────────────────────────────

# Сводная таблица: «Сумма по дебету/кредиту СЧЁТА» (итог по счёту) или «Остаток».
# ВАЖНО: «Сумма по дебету» БЕЗ слова «счёта» — это, наоборот, колонка суммы
# операции в ДЕТАЛЬНОЙ таблице (формат Сбербанка «счёт»), её отбрасывать нельзя.
_SUMMARY_HEADER_RE = re.compile(
    r"сумма\s+по\s+(?:деб\w*|кред\w*)\s+сч[её]т|остаток",
    re.IGNORECASE,
)


def is_summary_header(*texts: str) -> bool:
    """
    True если заголовок относится к СВОДНОЙ таблице («Сумма по … счёта»,
    «Остаток по счёту»), а не к детальной таблице операций («по дебету»/«по кредиту»
    или «Сумма по дебету» без слова «счёта»).
    Используется и для xlsx, и для docx/pdf (общий классификатор таблиц).
    """
    joined = " ".join(t for t in texts if t)
    return bool(_SUMMARY_HEADER_RE.search(joined))


def find_columns(sheet) -> tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    candidates = []
    for row in sheet.iter_rows():
        d_col = c_col = None
        d_text = c_text = ""
        for cell in row:
            if not (cell.value and isinstance(cell.value, str)):
                continue
            v = cell.value.lower()
            vs = v.strip()
            # Полные слова «дебит/дебет/кредит», а также аббревиатуры «Дт»/«Кт»
            # (формат ТБанк) — но аббревиатуры только как ЦЕЛАЯ ячейка, чтобы
            # не ловить их внутри других слов.
            if (("дебит" in v or "дебет" in v) or vs in ("дт", "д-т")) and d_col is None:
                d_col = cell.column; d_text = vs
            if ("кредит" in v or vs in ("кт", "к-т")) and c_col is None:
                c_col = cell.column; c_text = vs
        if d_col and c_col:
            candidates.append((row[0].row, d_col, c_col, is_summary_header(d_text, c_text)))

    if not candidates:
        return None, None, None, None

    # Отбрасываем СВОДНЫЕ таблицы («Сумма по … счёта»/«Остаток…»), если есть хотя
    # бы одна ДЕТАЛЬНАЯ. Порядок таблиц («Таблица 1/2») НЕ используем — он бывает
    # обратным (в части файлов детальная таблица идёт первой, сводная — второй).
    detail = [c for c in candidates if not c[3]]
    use = detail if detail else candidates
    triples = [(h, d, c) for (h, d, c, _s) in use]

    best_hrow, best_debit, best_credit = _pick_best_candidate(sheet, triples)

    date_col = _find_date_col_in_window(sheet, best_hrow, window=3)
    if date_col is None:
        date_col = _find_date_col_by_content(sheet, best_hrow, best_debit, best_credit)

    return best_hrow, date_col, best_debit, best_credit


def _is_valid_amount(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return 0 <= abs(value) <= 1e12
    if isinstance(value, str):
        v = to_float(value)
        if v is not None:
            return 0 <= abs(v) <= 1e12
    return False


def _pick_best_candidate(sheet, candidates: list) -> tuple:
    best = candidates[0]
    best_count = 0

    for hrow, d_col, c_col in candidates:
        count = 0
        for r in range(hrow + 1, sheet.max_row + 1):
            v_d = sheet.cell(row=r, column=d_col).value
            v_c = sheet.cell(row=r, column=c_col).value
            if _is_valid_amount(v_d) or _is_valid_amount(v_c):
                count += 1
            elif v_d is None and v_c is None and count > 0:
                empty_streak = sum(
                    1 for rr in range(r, min(r + 5, sheet.max_row + 1))
                    if sheet.cell(row=rr, column=d_col).value is None
                    and sheet.cell(row=rr, column=c_col).value is None
                )
                if empty_streak >= 5:
                    break
        if count > best_count:
            best_count = count
            best = (hrow, d_col, c_col)

    return best


def _find_date_col_in_window(sheet, header_row: int, window: int) -> Optional[int]:
    start = max(1, header_row - window)
    end = min(sheet.max_row, header_row + window)
    for r in range(start, end + 1):
        for cell in sheet[r]:
            if cell.value and isinstance(cell.value, str):
                if "дата" in cell.value.lower():
                    return cell.column
    return None


def _find_date_col_by_content(sheet, header_row: int,
                               debit_col: int, credit_col: int) -> Optional[int]:
    data_start = header_row + 1
    data_end = min(sheet.max_row, header_row + 35)
    left_bound = min(debit_col, credit_col)

    date_counts: dict[int, int] = {}
    total_counts: dict[int, int] = {}

    for r in range(data_start, data_end + 1):
        for c in range(1, left_bound):
            val = sheet.cell(row=r, column=c).value
            if val is None:
                continue
            total_counts[c] = total_counts.get(c, 0) + 1
            if parse_date(val) is not None:
                date_counts[c] = date_counts.get(c, 0) + 1

    best_col = None
    best_ratio = 0.5

    for col, count in date_counts.items():
        total = total_counts.get(col, 0)
        if total == 0:
            continue
        ratio = count / total
        if ratio > best_ratio:
            best_ratio = ratio
            best_col = col

    return best_col


# ── Вспомогательные функции ───────────────────────────────────────────────────

def is_column_numbering_row(debit_raw, credit_raw) -> bool:
    def small_int(v) -> bool:
        if isinstance(v, (int, float)):
            return float(v).is_integer() and 1 <= v <= 50
        if isinstance(v, str):
            s = v.strip()
            return s.isdigit() and 1 <= int(s) <= 50
        return False
    return small_int(debit_raw) and small_int(credit_raw)


def is_label(value) -> bool:
    if value is None or isinstance(value, (int, float)):
        return False
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not s or s in ("-", "—"):
        return False
    cleaned = s.replace("\xa0", "").replace(" ", "").replace(",", ".")
    if "-" in cleaned:
        parts = cleaned.split("-")
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return False
    try:
        float(cleaned)
        return False
    except ValueError:
        return True


def to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s or s in ("-", "—"):
            return None
        if "-" in s:
            parts = s.split("-")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                return float(f"{parts[0]}.{parts[1]}")
            return None
        cleaned = s.replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def get_description(sheet, row_idx: int, start_col: int) -> str:
    parts = []
    for col in range(start_col, sheet.max_column + 1):
        val = sheet.cell(row=row_idx, column=col).value
        if val is not None and isinstance(val, str) and val.strip():
            parts.append(val.strip())
    return " ".join(parts)


# Строки-итоги банка, у которых подпись стоит СЛЕВА от колонок сумм,
# а число — в самих колонках дебит/кредит. Их нельзя считать транзакциями.
_SUMMARY_ROW_RE = re.compile(
    r"^\s*(оборот|вход\w*\s+остаток|исход\w*\s+остаток|итого|остаток\s+по\s+сч)",
    re.IGNORECASE,
)


def is_summary_row(sheet, row_idx: int, left_bound: int) -> bool:
    """
    True если строка — итоговая/остаточная строка банка
    («Оборот», «Оборот за период», «Входящий/Исходящий остаток», «Итого»,
    «Остаток по счёту»), подпись которой находится в столбцах ЛЕВЕЕ сумм.
    У обычной транзакции слева стоит дата/номер, а не такое слово.
    """
    for col in range(1, max(1, left_bound)):
        v = sheet.cell(row=row_idx, column=col).value
        if isinstance(v, str) and _SUMMARY_ROW_RE.match(v.strip()):
            return True
    return False


# ── Обработка листа ───────────────────────────────────────────────────────────

def process_sheet(sheet) -> SheetResult:
    result = SheetResult(sheet_name=sheet.title)

    header_row, date_col, debit_col, credit_col = find_columns(sheet)
    if header_row is None:
        result.error = "Столбцы 'Дебит' и 'Кредит' не найдены"
        return result

    desc_start_col = max(date_col or 0, debit_col, credit_col) + 1
    data_start = header_row + 1

    if is_column_numbering_row(
        sheet.cell(row=data_start, column=debit_col).value,
        sheet.cell(row=data_start, column=credit_col).value,
    ):
        data_start += 1

    left_bound = min(debit_col, credit_col)

    for row_idx in range(data_start, sheet.max_row + 1):
        raw_d = sheet.cell(row=row_idx, column=debit_col).value
        raw_c = sheet.cell(row=row_idx, column=credit_col).value

        # Строка-итог банка («Оборот»/«Исходящий остаток»/«Итого»): подпись слева,
        # число в колонках сумм. Не транзакция — пропускаем (иначе перечёт ×2–3).
        if is_summary_row(sheet, row_idx, left_bound):
            result.skipped_rows += 1
            continue

        # Строка нумерации столбцов (1,2,…) — может встречаться и в середине листа
        # при нескольких таблицах-блоках.
        if is_column_numbering_row(raw_d, raw_c):
            result.skipped_rows += 1
            continue

        if is_label(raw_d) or is_label(raw_c):
            result.skipped_rows += 1
            continue

        d = to_float(raw_d)
        c = to_float(raw_c)

        # Значение > 10¹² — это не сумма операции, а номер счёта/реквизит,
        # ошибочно попавший в колонку (даёт мусор вроде дебета ~3·10¹⁹).
        if d is not None and abs(d) > 1e12:
            d = None
        if c is not None and abs(c) > 1e12:
            c = None

        if d is None and c is None:
            result.skipped_rows += 1
            continue

        # Оборот операций ДО исключения вкладов — для сверки с оборотом банка.
        result.raw_debit += d or 0.0
        result.raw_credit += c or 0.0

        description = get_description(sheet, row_idx, desc_start_col)
        if is_deposit_transfer(description):
            result.excluded_rows += 1
            continue

        tx_date = None
        if date_col:
            tx_date = parse_date(sheet.cell(row=row_idx, column=date_col).value)

        result.transactions.append(TransactionRow(
            date=tx_date,
            debit=d or 0.0,
            credit=c or 0.0,
            description=description,
            row_idx=row_idx,
        ))

    # Самопроверка: сверяем raw-оборот с задекларированным оборотом банка из файла.
    from app.core.reconcile import find_declared_turnover, reconcile
    result.declared_debit, result.declared_credit = find_declared_turnover(
        sheet, debit_col, credit_col)
    result.reconciled = reconcile(
        result.raw_debit, result.raw_credit,
        result.declared_debit, result.declared_credit)

    return result


# ── Обработка файла и папки ───────────────────────────────────────────────────

def process_file(filepath: str) -> AnalysisResult:
    kind = _detect_kind(filepath)

    if kind == "xml":
        from app.core.xml_reader import parse_xml_file
        return parse_xml_file(filepath)
    if kind == "docx":
        from app.core.docx_reader import parse_docx_file
        return parse_docx_file(filepath)
    if kind == "pdf":
        from app.core.pdf_reader import parse_pdf_file
        return parse_pdf_file(filepath)

    try:
        wb = _load_workbook(filepath)
    except Exception as e:
        raise RuntimeError(f"Не удалось открыть файл: {e}")

    result = AnalysisResult(filepath=filepath)
    for name in wb.sheetnames:
        result.sheets.append(process_sheet(wb[name]))
    return result


SUPPORTED_EXT = (".xlsx", ".xlsm", ".xls", ".xml", ".docx", ".doc", ".pdf")


def process_folder(folderpath: str) -> FolderResult:
    import os
    result = FolderResult(folderpath=folderpath)
    files = sorted([
        os.path.join(folderpath, f)
        for f in os.listdir(folderpath)
        if f.lower().endswith(SUPPORTED_EXT) and not f.startswith("~$")
    ])
    if not files:
        raise RuntimeError(
            "В папке не найдено поддерживаемых файлов (xlsx, xls, xml, docx, doc, pdf)"
        )
    for filepath in files:
        # Ошибка одного файла не должна ломать обработку всей папки: ловим её
        # и помечаем файл ошибкой чтения, продолжая с остальными.
        try:
            result.files.append(process_file(filepath))
        except Exception as e:
            ar = AnalysisResult(filepath=filepath, error=str(e))
            result.files.append(ar)
    # Фиксируем итоги БЕЗ дедупликации до того, как дедупликатор изменит
    # транзакции — чтобы показать оба итога (с дедупом / без) для прозрачности.
    result.raw_debit = sum(f.total_debit for f in result.files)
    result.raw_credit = sum(f.total_credit for f in result.files)
    deduplicate_folder(result)
    return result


# ── Определение типа файла ────────────────────────────────────────────────────

def _detect_kind(filepath: str) -> str:
    """
    Определяет тип файла по сигнатуре и расширению:
      'xml' | 'docx' | 'pdf' | 'table' (xlsx/xlsm/xls).
    Приоритет — сигнатуре (устойчиво к неверным расширениям).
    """
    ext = filepath.lower().rsplit(".", 1)[-1] if "." in filepath else ""
    try:
        with open(filepath, "rb") as f:
            head = f.read(8)
    except Exception:
        head = b""

    if ext == "xml" or head[:5] == b"<?xml":
        return "xml"
    if ext == "pdf" or head[:4] == b"%PDF":
        return "pdf"
    if head[:2] == b"PK":
        # OOXML: различаем Word (docx) и Excel (xlsx) по содержимому архива
        try:
            import zipfile
            with zipfile.ZipFile(filepath) as z:
                names = z.namelist()
                if any(n.startswith("word/") for n in names):
                    return "docx"
                if any(n.startswith("xl/") for n in names):
                    return "table"
        except Exception:
            pass
        return "table"
    if ext in ("docx", "doc"):
        # Не ZIP в начале (напр. старый бинарный .doc) — отдаём docx-ридеру,
        # он вернёт понятную ошибку обработки, а не исключение.
        return "docx"
    return "table"
