"""
Тесты логики обработки xlsx.
Запуск из корня проекта: python -m pytest tests/ -v
  или: python tests/test_processor.py
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import tempfile

from app.core.reader import process_file
from app.core.filters import is_deposit_transfer, is_interest

PASS = "\033[92m✓ PASS\033[0m"
FAIL = "\033[91m✗ FAIL\033[0m"
results = []


def make_wb(setup_fn):
    wb = openpyxl.Workbook()
    setup_fn(wb)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name

def check(name, filepath, sheet_idx, expected_debit, expected_credit,
          expected_excluded=None, expect_error=False):
    import traceback
    try:
        r = process_file(filepath)
        s = r.sheets[sheet_idx]
        if expect_error:
            ok = s.error is not None
            msg = f"error={s.error!r}"
        else:
            ok = abs(s.debit - expected_debit) < 0.01 and abs(s.credit - expected_credit) < 0.01
            if expected_excluded is not None:
                ok = ok and s.excluded_rows == expected_excluded
            msg = f"debit={s.debit:.2f} credit={s.credit:.2f} excl={s.excluded_rows} err={s.error}"
    except Exception as e:
        ok = False
        msg = f"EXCEPTION: {e}\n{traceback.format_exc()}"
    finally:
        try: os.unlink(filepath)
        except: pass

    results.append(ok)
    print(f"  {PASS if ok else FAIL} [{name}]")
    if not ok:
        print(f"         expected: D={expected_debit} C={expected_credit}")
        print(f"         got:      {msg}")


# ── Тесты фильтров ────────────────────────────────────────────────────────────

def test_filters():
    cases = [
        ("Перечисление средств во вклад",           True),
        ("Возврат депозита по договору 123",         True),
        ("Перечисление во вклад (депозит)",          True),
        ("Уплачены проценты по договору вклада",     False),
        ("Выплата %% по депозиту за период",         False),
        ("Выплата процентов по депозиту",            False),
        ("Начислены проценты по вкладу",             False),
        ("Оплата услуг ЖКХ",                        False),
        ("Пополнение р/счёта",                       False),
        ("",                                         False),
    ]
    all_ok = True
    for desc, expected in cases:
        got = is_deposit_transfer(desc)
        ok = got == expected
        all_ok = all_ok and ok
        if not ok:
            print(f"    FILTER FAIL: {desc!r} → got {got}, want {expected}")
    results.append(all_ok)
    print(f"  {PASS if all_ok else FAIL} [filters: is_deposit_transfer]")


# ── Тесты чтения xlsx ─────────────────────────────────────────────────────────

def t01(wb):
    ws = wb.active
    ws.append(["По дебиту счёта", "По кредиту счёта", "Назначение"])
    ws.append([1000, 0, "Оплата"]);  ws.append([0, 500, "Приход"]); ws.append([200, 0, "Аренда"])
check("01 базовый", make_wb(t01), 0, 1200, 500)

def t02(wb):
    ws = wb.active
    for _ in range(4): ws.append(["шапка"])
    ws.append(["По дебиту", "По кредиту", "Назначение"])
    ws.append([300, 0, "Платёж"]); ws.append([0, 150, "Возврат"])
check("02 заголовки с 5-й строки", make_wb(t02), 0, 300, 150)

def t03(wb):
    ws = wb.active
    for col, v in enumerate(["Банк","Дата","Документ","Плательщик","Получатель","По дебиту","По кредиту","Примечание"], 1):
        ws.cell(1, col, v)
    ws.cell(2,6,750); ws.cell(2,8,"Расход"); ws.cell(3,7,400); ws.cell(3,8,"Приход")
check("03 столбцы сдвинуты вправо", make_wb(t03), 0, 750, 400)

def t04(wb):
    ws = wb.active
    ws.append(["По дебету", "По кредиту"])
    ws.append(["1 234,56", None]); ws.append([None, "789,00"])
check("04 числа-строки русская локаль", make_wb(t04), 0, 1234.56, 789.00)

def t05(wb):
    ws = wb.active
    ws.append(["По дебиту", "По кредиту"])
    ws.append(["100000-50", None]); ws.append([None, "50000-25"])
check("05 формат '123456-78'", make_wb(t05), 0, 100000.50, 50000.25)

def t06(wb):
    # Сводная таблица (итоги «Сумма по … счёта») отбрасывается по содержанию,
    # берётся детальная — независимо от порядка таблиц на листе.
    ws = wb.active
    ws.append(["Таблица 1 (итоги)"])
    ws.append(["Сумма по дебету счета", "Сумма по кредиту счета"]); ws.append([9999, 8888]); ws.append([None, None])
    ws.append(["Таблица 2 (детали)"]); ws.append(["По дебиту", "По кредиту", "Назначение"])
    ws.append([100, None, "Расход"]); ws.append([None, 200, "Доход"])
check("06 сводная таблица отбрасывается, берётся детальная", make_wb(t06), 0, 100, 200)

def t07(wb):
    ws = wb.active
    ws.append(["По дебиту","По кредиту","Примечание"])
    ws.append([13, 14, 15]); ws.append([500, None, "Расход"]); ws.append([None, 300, "Доход"])
check("07 строка нумерации пропускается", make_wb(t07), 0, 500, 300)

def t08(wb):
    ws = wb.active
    ws.append(["По дебиту","По кредиту","Назначение"])
    ws.append([1000, None, "Перечисление средств во вклад"])
    ws.append([None, 500, "Поступление"]); ws.append([200, None, "Расход"])
check("08 исключение вклад", make_wb(t08), 0, 200, 500, expected_excluded=1)

def t09(wb):
    ws = wb.active
    ws.append(["По дебиту","По кредиту","Назначение"])
    ws.append([5000, None, "Перечисление средств во вклад (депозит)"])
    ws.append([None, 5000, "Возврат депозита по договору 123"])
    ws.append([None, 300, "Поступление"])
check("09 вклад+депозит оба исключаются", make_wb(t09), 0, 0, 300, expected_excluded=2)

def t10(wb):
    ws = wb.active
    ws.append(["По дебиту","По кредиту","Назначение"])
    ws.append([None, 1000, "Уплачены проценты по договору банковского вклада"])
    ws.append([None, 500,  "Выплата %% по вкладу за период"])
    ws.append([None, 200,  "Начислены проценты по депозиту"])
    ws.append([None, 800,  "Выплата процентов по депозиту 123"])
check("10 проценты НЕ исключаются", make_wb(t10), 0, 0, 2500, expected_excluded=0)

def t11(wb):
    ws = wb.active; ws.append(["Дата","Сумма"]); ws.append(["01.01.24", 100])
check("11 нет дебит/кредит → ошибка", make_wb(t11), 0, 0, 0, expect_error=True)

def t12(wb):
    ws = wb.active; ws.append(["По дебиту","По кредиту","Назначение"])
    ws.append([100, None, ""]); ws.append([None,None,None]); ws.append([None,None,None])
    ws.append([200, None, ""]); ws.append([None,None,None]); ws.append([None, 400, ""])
check("12 пустые строки внутри данных", make_wb(t12), 0, 300, 400)

def t13():
    def setup(wb):
        ws1 = wb.active; ws1.title = "Счёт 1"
        ws1.append(["По дебиту","По кредиту"]); ws1.append([1000,0]); ws1.append([0,600])
        ws2 = wb.create_sheet("Счёт 2")
        for _ in range(3): ws2.append(["шапка"])
        ws2.append(["Дебет","Кредит"]); ws2.append([400,None]); ws2.append([None,200])
    path = make_wb(setup)
    r = process_file(path)
    os.unlink(path)
    s1, s2 = r.sheets[0], r.sheets[1]
    ok = abs(s1.debit-1000)<0.01 and abs(s1.credit-600)<0.01 and abs(s2.debit-400)<0.01 and abs(s2.credit-200)<0.01
    results.append(ok)
    print(f"  {PASS if ok else FAIL} [13 два листа с разной структурой]")
t13()

def t14(wb):
    ws = wb.active; ws.append(["сумма по дебету счёта","сумма по кредиту счёта"])
    ws.append([777, None]); ws.append([None, 333])
check("14 'по дебету' через е", make_wb(t14), 0, 777, 333)

def t15(wb):
    ws = wb.active; ws.append(["ПО ДЕБИТУ СЧЁТА","ПО КРЕДИТУ СЧЁТА"])
    ws.append([888, None]); ws.append([None, 444])
check("15 заголовок ВЕРХНИЙ РЕГИСТР", make_wb(t15), 0, 888, 444)

def t16(wb):
    ws = wb.active; ws.append(["По дебиту","По кредиту","Назначение"])
    ws.append([500, 0, "Расход"]); ws.append([0, 300, "Приход"])
check("16 явные нули в ячейках", make_wb(t16), 0, 500, 300)

def t17(wb):
    ws = wb.active; ws.append(["По дебиту","По кредиту","Назначение"])
    ws.append([100, None, "Платёж"]); ws.append([None, 200, "Доход"])
    ws.append(["Итого", 100, 200])
check("17 строка 'Итого' пропускается", make_wb(t17), 0, 100, 200)

def t18(wb):
    # Формат банка с датой слева: строки-итоги «Оборот»/«остаток» стоят подписью
    # в столбце левее сумм — их нельзя считать транзакциями (F1).
    ws = wb.active
    ws.append(["Дата", "По дебиту", "По кредиту", "Назначение"])
    ws.append(["Входящий остаток", 0, 0])
    ws.append(["01.01.24", 0, 100, "Сбор с населения"])
    ws.append(["02.01.24", 0, 200, "Сбор с населения"])
    ws.append(["Оборот", 0, 300])
    ws.append(["Исходящий остаток", 0, 300])
check("18 строки-итоги Оборот/остаток не считаются (F1)", make_wb(t18), 0, 0, 300)

def t19(wb):
    # Несколько таблиц-блоков в одном листе: суммируем данные всех блоков,
    # строки-итоги и нумерацию пропускаем (F1).
    ws = wb.active
    ws.append(["Дата", "По дебиту", "По кредиту"])
    ws.append(["01.01.24", 100, 0])
    ws.append(["Оборот", 100, 0])
    ws.append(["Исходящий остаток", 100, 0])
    ws.append(["Дата", "По дебиту", "По кредиту"])
    ws.append(["02.01.24", 0, 200])
    ws.append(["Оборот", 0, 200])
check("19 несколько блоков: данные суммируем, итоги пропускаем (F1)", make_wb(t19), 0, 100, 200)

def t20(wb):
    # Значение >10¹² (номер счёта, попавший в колонку суммы) игнорируется (F4).
    ws = wb.active
    ws.append(["По дебиту", "По кредиту", "Назначение"])
    ws.append([30101810000000000805, None, "номер счёта как сумма"])
    ws.append([None, 500, "Реальный доход"])
check("20 значение >10¹² игнорируется (F4)", make_wb(t20), 0, 0, 500)

def t21(wb):
    # Короткая детальная «Таблица 2» выбирается вместо сводной «Таблица 1» (F4).
    ws = wb.active
    ws.append(["Таблица 1"])
    ws.append(["Остаток", "Сумма по дебету счета", "Сумма по кредиту счета"])
    ws.append([0, 0, 999999])
    ws.append(["Таблица 2"])
    ws.append(["По дебиту", "По кредиту", "Назначение"])
    ws.append([None, 999999, "единственная операция"])
check("21 короткая Таблица2 вместо сводной Таблица1 (F4)", make_wb(t21), 0, 0, 999999)

def t23(wb):
    # Формат ТБанк 2025: заголовки сумм — аббревиатуры «Дт»/«Кт» (F3).
    ws = wb.active
    ws.append(["Номер", "Дата", "Счет плательщика", "Дт", "Кт", "Назначение"])
    ws.append([1, "24.01.2025", "40703810100000733146", 0, 60000, "Взносы на капремонт"])
    ws.append([2, "27.02.2025", "40703810100000733146", 0, 40000, "Взносы на капремонт"])
check("23 формат ТБанк: заголовки Дт/Кт (F3)", make_wb(t23), 0, 0, 100000)

def t22_empty():
    # Пустой файл (0 байт) должен давать понятную ошибку, а не падать невнятно (F2).
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    raised = False
    try:
        process_file(tmp.name)
    except RuntimeError:
        raised = True
    except Exception:
        pass
    finally:
        try: os.unlink(tmp.name)
        except: pass
    results.append(raised)
    print(f"  {PASS if raised else FAIL} [22 пустой файл (0 байт) → понятная ошибка (F2)]")
t22_empty()

def t24_dedup_both():
    # Дедуп B: два перекрывающихся файла → доступны оба итога (без дедупа / с),
    # и дубликат по (дата, дебит, кредит) убирается ровно один раз.
    import tempfile, shutil
    from app.core.reader import process_folder
    d = tempfile.mkdtemp()
    try:
        for name, extra in (("a.xlsx", False), ("b.xlsx", True)):
            wb = openpyxl.Workbook(); ws = wb.active
            ws.append(["Дата", "По дебиту", "По кредиту", "Назначение"])
            ws.append(["01.01.24", 0, 100, "Платёж"])        # общая строка (дубль)
            if extra:
                ws.append(["02.01.24", 0, 50, "Платёж 2"])   # уникальная во 2-м файле
            wb.save(os.path.join(d, name))
        r = process_folder(d)
        raw = r.raw_total_credit    # без дедупа: 100 + (100+50) = 250
        ded = r.total_credit        # с дедупом:  100 + 50       = 150
        ok = abs(raw - 250) < 0.01 and abs(ded - 150) < 0.01 and r.total_deduplicated == 1
        results.append(ok)
        print(f"  {PASS if ok else FAIL} [24 дедуп: без={raw:.0f} с дедупом={ded:.0f} убрано={r.total_deduplicated}]")
    finally:
        shutil.rmtree(d, ignore_errors=True)
t24_dedup_both()

def t25_grid():
    # GridSheet должен давать тот же результат, что openpyxl-лист на тех же данных
    # (адаптер для docx/pdf, чтобы переиспользовать process_sheet без изменений).
    from app.core.grid import GridSheet
    from app.core.reader import process_sheet
    data = [
        ["Дата", "По дебиту", "По кредиту", "Назначение"],
        ["01.01.24", 0, 100, "Сбор"],
        ["02.01.24", 50, 0, "Оплата"],
        ["Оборот", 50, 100],   # строка-итог, должна пропускаться
    ]
    wb = openpyxl.Workbook(); ws = wb.active
    for row in data:
        ws.append(row)
    r_xlsx = process_sheet(ws)
    r_grid = process_sheet(GridSheet(data, "T"))
    ok = (abs(r_xlsx.debit - r_grid.debit) < 0.01 and abs(r_xlsx.credit - r_grid.credit) < 0.01
          and abs(r_grid.debit - 50) < 0.01 and abs(r_grid.credit - 100) < 0.01)
    results.append(ok)
    print(f"  {PASS if ok else FAIL} [25 GridSheet == openpyxl: D={r_grid.debit:.0f} K={r_grid.credit:.0f}]")
t25_grid()

def t26_docx():
    # Сквозной тест docx-пути: диспетчеризация → docx_reader → GridSheet → process_sheet.
    from docx import Document
    doc = Document()
    t = doc.add_table(rows=0, cols=4)
    for row in [["Дата", "По дебиту", "По кредиту", "Назначение"],
                ["01.01.24", "0", "100", "Сбор"],
                ["02.01.24", "50", "0", "Оплата"],
                ["Оборот", "50", "100", ""]]:  # строка-итог должна пропускаться
        cells = t.add_row().cells
        for i, v in enumerate(row):
            cells[i].text = v
    tmp = tempfile.NamedTemporaryFile(suffix=".docx", delete=False); tmp.close()
    doc.save(tmp.name)
    try:
        r = process_file(tmp.name)
        ok = abs(r.total_debit - 50) < 0.01 and abs(r.total_credit - 100) < 0.01
        msg = f"D={r.total_debit:.0f} K={r.total_credit:.0f}"
    except Exception as e:
        ok = False; msg = f"EXC {e}"
    finally:
        try: os.unlink(tmp.name)
        except: pass
    results.append(ok)
    print(f"  {PASS if ok else FAIL} [26 docx сквозной: {msg}]")
t26_docx()

def t27_pdf_robust():
    # Битый PDF → штатная ошибка обработки, БЕЗ падения (устойчивость).
    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    tmp.write(b"this is not a real pdf"); tmp.close()
    ok = False
    try:
        r = process_file(tmp.name)
        ok = bool(r.sheets) and all(s.error for s in r.sheets)
    except Exception:
        ok = False
    finally:
        try: os.unlink(tmp.name)
        except: pass
    results.append(ok)
    print(f"  {PASS if ok else FAIL} [27 битый PDF → ошибка, не краш (устойчивость)]")
t27_pdf_robust()

def t28_resilience():
    # Ошибка одного файла не должна ломать обработку всей папки (n файлов, m с ошибкой).
    from app.core.reader import process_folder
    import shutil
    d = tempfile.mkdtemp()
    try:
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Дата", "По дебиту", "По кредиту", "Назначение"])
        ws.append(["01.01.24", 0, 100, "Сбор"])
        wb.save(os.path.join(d, "good.xlsx"))
        open(os.path.join(d, "bad.xlsx"), "wb").close()  # 0 байт — битый
        r = process_folder(d)  # НЕ должно бросить исключение
        by = {f.filename: f for f in r.files}
        ok = (len(r.files) == 2
              and by["good.xlsx"].status == "ok" and abs(by["good.xlsx"].total_credit - 100) < 0.01
              and by["bad.xlsx"].status == "read_error")
        msg = f"good={by['good.xlsx'].status} bad={by['bad.xlsx'].status}"
    except Exception as e:
        ok = False; msg = f"EXC {e}"
    finally:
        shutil.rmtree(d, ignore_errors=True)
    results.append(ok)
    print(f"  {PASS if ok else FAIL} [28 устойчивость: {msg}]")
t28_resilience()

def t29_reconcile():
    # Самопроверка: сверка расчёта с «Оборот»-строкой банка.
    from app.core.reader import process_sheet
    from app.core.grid import GridSheet
    ok_true = process_sheet(GridSheet([
        ["Дата", "По дебиту", "По кредиту", "Назначение"],
        ["01.01.24", "0", "100", "Сбор"],
        ["02.01.24", "0", "200", "Сбор"],
        ["Оборот", "0", "300"]])).reconciled is True
    ok_false = process_sheet(GridSheet([
        ["Дата", "По дебиту", "По кредиту", "Назначение"],
        ["01.01.24", "0", "100", "Сбор"],
        ["Оборот", "0", "999"]])).reconciled is False
    ok_none = process_sheet(GridSheet([
        ["Дата", "По дебиту", "По кредиту", "Назначение"],
        ["01.01.24", "0", "100", "Сбор"]])).reconciled is None
    ok = ok_true and ok_false and ok_none
    results.append(ok)
    print(f"  {PASS if ok else FAIL} [29 самопроверка: совпало={ok_true} не_совпало={ok_false} нет_сверки={ok_none}]")
t29_reconcile()

test_filters()

# ── Итог ──────────────────────────────────────────────────────────────────────
total = len(results); passed = sum(results)
print(f"\n{'='*50}")
print(f"Результат: {passed}/{total} тестов пройдено")
if passed < total:
    print("\033[91mЕсть провалившиеся тесты!\033[0m")
else:
    print("\033[92mВсе тесты пройдены ✓\033[0m")
